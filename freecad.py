# Step file ops
import members
import random
import colormaps
from numpy import linspace

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd


# make allpoint
# p1 = FreeCAD.Vector(1000, 0, 0)

# importHeader = [['import Draft, Arch']]

def pointCloud(reg):
    lst = []
    for i, point in enumerate(reg['Point']):
        lst.append(['p' + point.name + ' = FreeCAD.Vector' + '(' + str(round(point.x * 1000)) +
                    ',' + str(round(point.y * 1000)) + ',' + str(round(point.z * 1000)) + ')'])
    return lst


def beamGen(reg, sections=False, colour='Grey'):
    # sections = 50 makes all pipes with 50 diam
    lst = []

    for i, beam in enumerate(reg['Beam']):
        lst.append(['Line' + str(i) + ' = Draft.makeWire([' + 'p' + beam.start.name + ',p' + beam.end.name + '])'])
        if isinstance(sections, bool):
            if sections:
                if isinstance(beam.section, members.Tube):
                    lst.append(['Arch.makePipe(Line' + str(i) + ',' + str(
                        round(beam.section.diam * 10, 2)) + ').WallThickness=' + str(
                        round(beam.section.thick * 10, 2))])
        # General pipe section for all beams
        elif isinstance(sections, list):
            lst.append(['Arch.makePipe(Line' + str(i) + ',' + str(sections[0]) + ')'])
        if colour is not False:
            lst.append([f'Pipe.ViewObject.ShapeColor = (128,128,128)'])
            lst.append([f'Pipe.ViewObject.LineColor = (60, 60, 60)'])
        return lst


def beamgroup(reg, sections=True, colour_opt=False, mix=False, export=False, groups=True):
    lst = []
    # colour = (120, 120, 120)
    #group_selection = ['BR1', 'HSP', 'JLT', 'LG1', 'LTR', 'PAD', 'SLE', 'VP', 'YK1', 'YK2', 'YK3', 'SLG','WOO']
    group_selection = ['BR1', 'HSP', 'JLT', 'LG1', 'LTR', 'PAD', 'SLE', 'VP', 'YK1', 'YK2', 'YK3']
    if not groups:
        group_selection = [x.name for x in reg['Group']]

    def coloursel(n, total=len(group_selection), colourmap=colormaps.Spectral, invert=False):
        # print(int(n / total * len(colourmap)))
        #
        sequence = linspace(0, len(colourmap) - 1, total)
        if invert:
            colourmap = colourmap[::-1]

        expanded_seq = [int(value) for value in sequence]
        print(len(colourmap), n, expanded_seq[n])

        return colourmap[expanded_seq[n]]

    if mix:
        size = len(group_selection)
        sequence = random.sample(list(range(size)), size)
        print(sequence)

    data = []

    for i, group in enumerate(reg['Group']):
        if group.name in group_selection:

            if colour_opt :
                if mix:
                    colour = coloursel(sequence[group_selection.index(group.name)])
                    print ('C1',colour)
                else:
                    colour = coloursel(group_selection.index(group.name))
                    print('C2',colour)

            elif colour_opt is False:  # Grey
                colour = (128, 128, 128)

            beam0 = group.elem[0]

            if isinstance(beam0.section, members.Tube):
                data.append([group.name, 'Group Name', 'Material', beam0.section.diam, beam0.section.thick, colour])
            # data.append([group.name, 'Group Name', 'Material', beam0.section.name, beam0.section.thick, colour])

            for j, beam in enumerate(group.elem):
                if isinstance(beam.section, members.Tube):

                    lst.append(
                        ['Line0' + str(i) + str(
                            j) + ' = Draft.makeWire([' + 'p' + beam.start.name + ',p' + beam.end.name + '])'])
                    if sections:
                        diameter = round(beam.section.diam * 10, 2)
                    else:
                        diameter = 30

                    lst.append(['Pipe=Arch.makePipe(Line0' + str(i) + str(j) + ',' + str(diameter) + ')'])

                    if sections:
                        lst.append([f'Pipe.WallThickness={str(round(beam.section.thick * 10, 2))}'])

                        lst.append([f'Pipe.ViewObject.ShapeColor = {colour}'])

                elif isinstance(beam.section, members.FlatBar):  # EDITING
                    lst.append(
                        ['Line0' + str(i) + str(
                            j) + ' = Draft.makeWire([' + 'p' + beam.start.name + ',p' + beam.end.name + '])'])

                    # if sections:

                    diameter = 30

                    lst.append(['Pipe=Arch.makePipe(Line0' + str(i) + str(j) + ',' + str(diameter) + ')'])

                    if sections:
                        # make flatbar

                        lst.append([f"section = Arch.makeProfile(profile=[0, 'REC', '{beam.section.name}', 'R',"
                                    f" {beam.section.th * 10}, {beam.section.height * 10}])"])
                        # makeProfile(profile=[0, 'REC', 'REC100x100', 'R', 100, 100]):
                        # lst.append(f'Pipe.Section'])
                        lst.append([f'Pipe.Profile=section'])
                        lst.append([f'Pipe.ViewObject.ShapeColor = {colour}'])

                elif isinstance(beam.section, members.Isection):  # EDITING


                    #members.Isection(name, height, width, th_f, th_w)
                    lst.append(
                        ['Line0' + str(i) + str(
                            j) + ' = Draft.makeWire([' + 'p' + beam.start.name + ',p' + beam.end.name + '])'])

                    # if sections:

                    diameter = 30

                    lst.append(['Pipe=Arch.makePipe(Line0' + str(i) + str(j) + ',' + str(diameter) + ')'])

                    if sections:
                        # make Isection

                        lst.append([f"section = Arch.makeProfile(profile=[0, 'I', '{beam.section.name}', 'H',"
                                    f"{beam.section.height * 10},{beam.section.width*10},{beam.section.tw*10},"
                                    f"{beam.section.tf*10}])"])
                        # makeProfile(profile=[0, 'REC', 'REC100x100', 'R', 100, 100]):
                        # lst.append(f'Pipe.Section'])
                        lst.append([f'Pipe.Profile=section'])
                        lst.append([f'Pipe.ViewObject.ShapeColor = {colour}'])

                elif isinstance(beam.section, members.RecSection):  # EDITING


                    #members.Isection(name, height, width, th_f, th_w)
                    lst.append(
                        ['Line0' + str(i) + str(
                            j) + ' = Draft.makeWire([' + 'p' + beam.start.name + ',p' + beam.end.name + '])'])

                    # if sections:

                    diameter = 30

                    lst.append(['Pipe=Arch.makePipe(Line0' + str(i) + str(j) + ',' + str(diameter) + ')'])

                    if sections:
                        # make Isection

                        lst.append([f"section = Arch.makeProfile(profile=[0, 'RHS', '{beam.section.name}', 'RH',"
                                    f"{beam.section.height * 10},{beam.section.width*10},{beam.section.th*10}])"])
                        # makeProfile(profile=[0, 'REC', 'REC100x100', 'R', 100, 100]):
                        # lst.append(f'Pipe.Section'])
                        lst.append([f'Pipe.Profile=section'])
                        lst.append([f'Pipe.ViewObject.ShapeColor = {colour}'])


    if export:
        export_list_to_excel(data, 'Section_Colors.xlsx')

    return lst

    # read from excel or txt - create folder with joints
    # export to abaqus #TODO MAYBE


def jointGen(reg, nodelst=None, inputfile='xls', sections=False, stub_len=2, folder=False):
    lst = []
    print('Stub len', stub_len)
    # INPUT TYPE - xls,or txt
    if inputfile == 'xls':
        # get from excel
        print('From excel')
    else:
        # get input from txt,csv
        print('Nodes from txt')
    # If not joints specified populates with all joints
    if nodelst is None:
        nodelst = []
        for joint in reg['Joint']:
            nodelst.append(joint.point.name)
    # DO for selected joints
    i = len(members.reg['Beam']) + 1
    # print('entrou')

    for node_name in nodelst:
        print('Generate Node', node_name)

        for joint in reg['Joint']:
            # print(node_name,joint.point.name)

            if joint.point.name == node_name:
                print(joint.point.name)
                # print(joint.name)
                for beam in joint.beams:
                    # Check which stub length to use
                    if hasattr(beam.section, "diam"):
                        print('stub check')
                        # print(stub_len*beam.section.diam*0.01,beam.length())

                        # stub =stub_len * beam.section.diam * 0.01
                        stub = min(beam.length(), 0.8)
                        print('stub used', stub)
                        if beam.start == joint.point:
                            vec = members.Vector(beam.start, beam.end)
                        else:
                            vec = members.Vector(beam.end, beam.start)
                        # print('vec',vec.dim)
                        pivot = vec.node_along(point='A', dist=stub)  # TODO not just for tubular
                        # print('BeamA',beam.start.pos)
                        # print('pivot',pivot.pos)
                        # print('thickness',beam.section.thick)

                        auxi = f'FreeCAD.Vector{tuple([round(x * 1000) for x in pivot.pos])}'
                        lst.append([f'Line{i} = Draft.makeWire([p{joint.point.name},{auxi}])'])
                        _str = f'Wire = Arch.makePipe(Line{i},{round(beam.section.diam * 10)},' \
                               f'{beam.section.thick * 10})'
                        _str += f'.WallThickness={beam.section.thick * 10}'
                        lst.append([_str])
                    elif hasattr(beam.section, "height"):
                        stub = str(stub_len * beam.section.height)
                    else:
                        stub = str(300)
                        print(f'using default stub for {beam.section.name}')
                    if folder:
                        lst.append([f'folder.addObject(Wire)'])
                    i += 1
                break
    # ('aqui', lst)
    return lst

    # Line = Draft.makeWire([p1, p2, p3, p4])


def jointSpheres(reg, nodelst=None, diam=200, colour='saipem'):
    # nodelst=[]
    saipem = (0, 85, 127)
    transparency = 50
    lst = ['doc=FreeCAD.ActiveDocument']
    if nodelst is None:
        print('here')
        for i, joint in enumerate(members.reg['Joint']):
            if i > 5:
                return lst
            else:
                position = tuple([xyz * 1000 for xyz in joint.point.pos])
                lst.append([f'Sphere{i} = Part.makeSphere({diam * 0.5},FreeCAD.Vector{position})'])
                lst.append('sphere = doc.addObject("Part::Feature", "Sphere")')
                lst.append(f'sphere.Shape = Sphere{i}')
                lst.append(f'sphere.ViewObject.ShapeColor = {saipem}')
                lst.append(f'sphere.ViewObject.Transparency = {transparency}')

                # print(lst)

    return lst


def jointLabels(reg, nodelst=None, delta_text=[0.5, 0, 1], sections=False):
    # nodelst=[]
    saipem = (0, 85, 127)
    white = (255, 255, 255)
    black = (0, 0, 0)
    colour = saipem
    transparency = 50
    lst = ['doc=FreeCAD.ActiveDocument']

    default_block = f'''
label.ViewObject.FontSize = 350  # Set the default font size (adjust as needed)
label.ViewObject.ArrowSize = 100  # Set the default arrow size (adjust as needed)
label.ViewObject.ArrowType = "Dot"  # Set the default arrow type (adjust as needed)
label.ViewObject.TextColor = {black}  # Set the default text color (black)
label.ViewObject.LineColor = {black}  # Set the default line color (black)
label.ViewObject.FontName = "Arial"  # Set the default font name (adjust as needed)
label.ViewObject.Frame = 'Rectangle' '''


# CASES are generated by test.bat from analysis directory


    # cases = {
    #     '1000': ['0170', '0089', '0053', '0164', '0134'],
    #     '2000': ['0240', '0127', '0057', '0058', '0048'],
    #     '3100': ['0048', '0055', '0056', '0057', '0058'],
    #     '3200': ['0048', '0055', '0056', '0057', '0058'],
    #     '3300': ['0048', '0055', '0056', '0057', '0058'],
    #     '3600': ['0048', '0055', '0056', '0057', '0058'],
    #     '4000': ['0170', '0089', '0053', '0164', '0134']}

    cases = {
        '6200': ['0161', '0134', '0253', '0055', '0056']}

    # cases = {'5000': ['0008', '0054', '0122', '0121', '0041'],
    #         '5200': ['0172', '0107', '0108', '0080', '0008']}

    rotation = [f'rotation = FreeCAD.Rotation(FreeCAD.Vector(1.00, 0.00, 0.00), 90.00)']

    for case in cases:
        lst.append(f'\nfolder = doc.addObject("App::DocumentObjectGroup","CASE {case}")')
        for node in cases[case]:
            print('Processing Node - ' + node)
            node = members.get_obj(node, 'Point')
            text_position = tuple([(c1 + c2) * 1000 for c1, c2 in zip(node.pos, delta_text)])
            lst.append([f'position = FreeCAD.Vector{node * 1000}'])
            lst.append([f'txt_position = FreeCAD.Vector({text_position})'])
            lst.append(rotation)
            lst.append([f'txt_placement = FreeCAD.Placement(txt_position, rotation)'])
            lst.append([f'label = Draft.make_label(position, txt_placement, custom_text="{node.name}", distance=-100)'])
            lst.append(default_block)
            lst.append(f'folder.addObject(label)')
        # lst.append(f'folder.addObject(label)')
        if sections:
            print('Sections')
            lst += jointGen(reg, cases[case], inputfile='xls', sections=True, stub_len=1, folder=False)

            # print(lst)
    return lst


def extendPoint(pointA, pointB, dist=0, mode=1):
    # modetypes=['fromA','fromB','middle']
    vec = members.Vector(pointA, pointB)
    if mode == 1:
        point = 'A'
    elif mode == 2:
        point = 'B'
    elif mode == 'middle':  # TODO finish if necessary
        point = 'middle'

    return vec.node_along(point=point, dist=dist)


import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def rgb_to_hex(rgb):
    return '%02x%02x%02x' % rgb


def export_list_to_excel(data_list, output_filename):
    # Create a DataFrame from the list
    df = pd.DataFrame(data_list,
                      columns=['Group', 'Group Element', 'Material', 'Diameter', 'Thickness', 'Colour Code'])

    # Convert the RGB code to a string representation
    df['Colour Code'] = df['Colour Code'].apply(rgb_to_hex)

    # Create an Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the DataFrame to the sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Apply fill color to the last cell of each row
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.max_column, max_col=sheet.max_column):
        cell = row[0]
        rgb_code = cell.value
        fill = openpyxl.styles.PatternFill(start_color=rgb_code, end_color=rgb_code, fill_type="solid")
        cell.fill = fill

    # Save the workbook
    workbook.save(output_filename)
