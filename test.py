import core
import os
# import pandas as pd
from pprint import pprint as pp

# aux = list()
# for i in [1,2,3,4,3,4,5]:
#    if i not in aux:
#        aux.append(i)
#        print (aux)

# OBTER OUPUT DE CODEX PARA PARSING
string = [0] * 4
string[
    0] = 'FA1 FA01-FA02 7002  12.09 110.3   4594.3 2588.0 2340.0   .3E+03 .8E+05 .3E+03 .3E+03   TN+BN    0.33   0.33 ' \
         '' \
         '' \
         '' \
         '' \
         '' \
         '' \
         '' \
         '' \
         '  0.85   0.85'
string[1] = 'SECT 22X22     TUB                               55.8802.223'
string[2] = 'SECT LPD       PRI                               100.00     10.000'
string[3] = 'SECT 28X23     TUB                               71.1202.382'


def posCount(string):
    aux = []
    aux2 = []
    for pos, letter in enumerate(string):
        aux.append(pos)
        aux2.append(letter)
    print([aux, aux2])

    # pp(aux2)


# posCount(string[3])

def stringCount(string, sep=False):
    aux = []
    words = string.split()
    for word in words:
        if sep is False or word.count('.') < 2:
            start = string.find(word)
            end = start + len(word)
            aux = aux + [[start, end]]
        else:
            print('este')
            start = string.find(word)
            end = start + len(word)
            aux = aux + [[start, end]]
    print(words)
    # print(core.Color.OKGREEN,aux,core.Color.ENDC)
    core.Color.prtOKRANDOM(aux)
    return aux

stringCount('JOINT 0189      7.    -2.     0. 63.500       -17.54')

# Section generator
file = os.getcwd() + os.sep + 'sect.txt'


def textCount(filePath):
    text = [line for line in open(filePath) if len(line.split()) > 3]
    media = []
    maxid = []
    for line in text:
        media.append(stringCount(line))

    minwords = min(len(line) for line in media)


# textCount(file)

# Section generator
# file = core.SACSdir + os.sep + 'Sections.txt'
# text = [line.rstrip('\n').split() for line in open(file)]
# pp(text)
# for line in text:
#     diam = round(float(line[1])/10,2)
#     thick = round(float(line[2])/10,2)
#     print ('SECT ',' '*(7-len(line[0])),line[0],'   TUB',' '*31,
#            ' '*(5-len(str(diam))),diam,' '*(5-len(str(thick))),thick,sep='')

# rgb = []
# hexlist = ['#440154FF', '#481567FF', '#482677FF', '#453781FF', '#404788FF',
#            '#39568CFF', '#33638DFF', '#2D708EFF', '#287D8EFF', '#238A8DFF',
#            '#1F968BFF', '#20A387FF', '#29AF7FFF', '#3CBB75FF', '#55C667FF',
#            '#73D055FF', '#95D840FF', '#B8DE29FF', '#DCE319FF', '#FDE725FF']
# for hexc in hexlist:
#     h = hexc.lstrip('#')
#     val = tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))
#     print('RGB =', val)
#     rgb.append(val)
#
# print(rgb)

import pandas as pd

# Read the Excel file

for excelname in os.listdir(os.getcwd() + os.sep + 'GEN'):
    if excelname.endswith(".xlsx"):
        break

df = pd.read_excel(os.getcwd() + os.sep + 'GEN' + os.sep + excelname, sheet_name='Global Status',
                   header=None)


# Define the column number (e.g., column A is 0, column B is 1, and so on)

# Iterate through the rows
def iterator(df, col_num):
    node_lists = []
    current_list = []
    for index, row in df.iterrows():
        if pd.isnull(row[col_num]):
            # Empty row encountered, append current list and start a new one
            if current_list:
                node_lists.append(current_list)

                current_list = []
        else:
            # Non-empty row, add node to the current list
            current_list.append(row[col_num])

    # Append the last list if it is not empty
    if current_list:
        node_lists.append(current_list)
    #print (node_lists)
    return node_lists

# Print screenshot sequence
print('\n\nd = 280')
for seq, seq2 in zip(iterator(df, 7), iterator(df, 6)):
    #print (seq2)
    seq[0] = str(seq2[1])[:2] + '00prt'
    print(f"whole('{seq[0]}',{seq[1:]}, 370, d, 80)\nd += 24")

#Print NODE freecad data
for seq, seq2 in zip(iterator(df, 7), iterator(df, 6)):
    #print (seq2)
    seq[0] = str(seq2[1])[:2] + '00'
    print(f"'{seq[0]}':{seq[1:]},")

list_of_colours = ['Accent', 'Accent_r', 'Blues', 'Blues_r', 'BrBG', 'BrBG_r', 'BuGn', 'BuGn_r', 'BuPu', 'BuPu_r',
                   'CMRmap', 'CMRmap_r', 'Dark2', 'Dark2_r', 'GnBu', 'GnBu_r', 'Greens', 'Greens_r', 'Greys', 'Greys_r',
                   'OrRd', 'OrRd_r',
                   'Oranges', 'Oranges_r', 'PRGn', 'PRGn_r', 'Paired', 'Paired_r', 'Pastel1', 'Pastel1_r', 'Pastel2',
                   'Pastel2_r',
                   'PiYG', 'PiYG_r', 'PuBu', 'PuBuGn', 'PuBuGn_r', 'PuBu_r', 'PuOr', 'PuOr_r', 'PuRd', 'PuRd_r',
                   'Purples',
                   'Purples_r', 'RdBu', 'RdBu_r', 'RdGy', 'RdGy_r', 'RdPu', 'RdPu_r', 'RdYlBu', 'RdYlBu_r', 'RdYlGn',
                   'RdYlGn_r',
                   'Reds', 'Reds_r', 'Set1', 'Set1_r', 'Set2', 'Set2_r', 'Set3', 'Set3_r', 'Spectral', 'Spectral_r',
                   'Wistia',
                   'Wistia_r', 'YlGn', 'YlGnBu', 'YlGnBu_r', 'YlGn_r', 'YlOrBr', 'YlOrBr_r', 'YlOrRd', 'YlOrRd_r',
                   'afmhot',
                   'afmhot_r', 'autumn', 'autumn_r', 'binary', 'binary_r', 'bone', 'bone_r', 'brg', 'brg_r', 'bwr',
                   'bwr_r',
                   'cividis', 'cividis_r', 'cool', 'cool_r', 'coolwarm', 'coolwarm_r', 'copper', 'copper_r',
                   'cubehelix',
                   'cubehelix_r', 'flag', 'flag_r', 'gist_earth', 'gist_earth_r', 'gist_gray', 'gist_gray_r',
                   'gist_heat',
                   'gist_heat_r', 'gist_ncar', 'gist_ncar_r', 'gist_rainbow', 'gist_rainbow_r', 'gist_stern',
                   'gist_stern_r',
                   'gist_yarg', 'gist_yarg_r', 'gnuplot', 'gnuplot2', 'gnuplot2_r', 'gnuplot_r', 'gray', 'gray_r',
                   'hot', 'hot_r',
                   'hsv', 'hsv_r', 'inferno', 'inferno_r', 'jet', 'jet_r', 'magma', 'magma_r', 'nipy_spectral',
                   'nipy_spectral_r',
                   'ocean', 'ocean_r', 'pink', 'pink_r', 'plasma', 'plasma_r', 'prism', 'prism_r', 'rainbow',
                   'rainbow_r', 'seismic',
                   'seismic_r', 'spring', 'spring_r', 'summer', 'summer_r', 'tab10', 'tab10_r', 'tab20', 'tab20_r',
                   'tab20b',
                   'tab20b_r', 'tab20c', 'tab20c_r', 'terrain', 'terrain_r', 'turbo', 'turbo_r', 'twilight',
                   'twilight_r',
                   'twilight_shifted', 'twilight_shifted_r', 'viridis', 'viridis_r', 'winter', 'winter_r']

import matplotlib.pyplot as plt


def colorwheel(colormap, num_colors):
    # cmap = cm.get_cmap(colormap)
    cmap = plt.get_cmap(colormap)
    step = max(int(cmap.N / num_colors), 1)
    aux = [tuple(int(255 * x) for x in cmap(i)[:3]) for i in range(0, cmap.N, step)]
    print(f'{colormap} = {aux}')


#for color in list_of_colours:
#    colorwheel(color, 100)

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import colormaps


def change_cell_colors(rgb_list, output_file):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Iterate over the RGB tuples and set cell colors
    for i, rgb in enumerate(rgb_list, start=1):
        fill = PatternFill(start_color=f"FF{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}", fill_type="solid")
        sheet.cell(row=i, column=1).fill = fill

    # Save the workbook
    workbook.save(output_file)


# Example usage
rgb_list = colormaps.turbo
output_file = "colored_cells.xlsx"

#change_cell_colors(rgb_list, output_file)
